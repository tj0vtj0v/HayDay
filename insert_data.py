from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from string import ascii_uppercase

from DBConnection import DbConnection

# establish DB connection
db = DbConnection('hayday', 'i_am_hayday', 'hayday')

# connect EXCEL file
wb: Workbook
wb = load_workbook('./Data/DB_Layout.xlsx')

ws_source = wb['Source']
ws_ingredients = wb['Ingredients']
ws_items = wb['Items']


def transfer():
    def get_values(ws: Worksheet, row: int, cols: int, fill_value: str = 'NULL') -> str:
        value = ''
        for col in range(0, cols):
            value += '"' + str(ws[f'{ascii_uppercase[col]}{row}'].value) + '", '

        return value.removesuffix(", ").replace('""', fill_value).replace('"None"', fill_value)

    row = 2
    while ws_source[f'A{row}'].value is not None:
        db.make_entry('source', get_values(ws_source, row, 3))
        row += 1

    row = 2
    while ws_ingredients[f'A{row}'].value is not None:
        db.make_entry('ingredients', get_values(ws_ingredients, row, 9))
        row += 1

    row = 2
    while ws_items[f'A{row}'].value is not None:
        db.make_entry('items', get_values(ws_items, row, 10, '0'))
        row += 1


def modify():
    # set production time of mine products to 0
    db.exec_statement('''UPDATE items SET `production_time` = 0, `mastered_time` = 0 WHERE id <= 198 AND id >= 193;''')

    # update false production times, order by index
    # lobster tail?
    db.exec_statement('''UPDATE items SET `production_time` = 25 WHERE name = "Winter Veggies"''')
    db.exec_statement('''UPDATE items SET `production_time` = 30 WHERE name = "Gingerbread Cookie"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 76 WHERE name = "Honey Mask"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 63 WHERE name = "Pineapple Cake"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 63 WHERE name = "Onion Dog"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 114 WHERE name = "Mint Ice Cream"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 331 WHERE name = "Grape Jam"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 187 WHERE name = "Potted Plant"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 204 WHERE name = "Pickles"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 29 WHERE name = "Cucumber Sandwich"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 29 WHERE name = "Plum Smoothie"''')
    db.exec_statement('''UPDATE items SET `production_time` = 75 WHERE name = "Broccoli Soup"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 63 WHERE name = "Nachos"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 34 WHERE name = "Honey Tea"''')
    db.exec_statement('''UPDATE items SET `mastered_time` = 29 WHERE name = "Mint Tea"''')


transfer()
modify()
